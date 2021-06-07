from openpyxl import Workbook, load_workbook

wb = Workbook()
sheet = wb.active

data = (
    (11, 48, 50),
    (81, 30, 82),
    (20, 51, 72),
    (21, 14, 60),
    (28, 41, 49),
    (74, 65, 53),
    ("Peter", 'Andrew', 45.63)
)

for i in data:
    sheet.append(i)
wb.save('append.xlsx')


# read data to append.xlsx
# wb = load_workbook('append.xlsx')
# sheet = wb.active
#
# max_column = sheet.max_column
# max_row = sheet.max_row
#
# for row in sheet.iter_rows(min_row=1, min_col=1, max_row=max_row, max_col=max_column):
#     for cell in row:
#         print(cell.value, end=" ")
#     print()
